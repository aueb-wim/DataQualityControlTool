# cde.py
import re

from openpyxl import load_workbook

from mipqctool.helpers import edit_distance_f1
from mipqctool.exceptions import CdeDictError
from mipqctool.model.mapping.functions import Replacement
from mipqctool.config import LOGGER

QC_FROM_CDE_NUMERICAL = ['numerical', 'numeric', 'real']
QC_FROM_CDE_INTEGER = ['integer', 'int']
QC_FROM_CDE_NOMINAL = ['nominal', 'ordinal', 'binomial', 'polynomial']
QC_FROM_CDE_DATE = ['date']


class CdeDict(object):
    """Class to store information about cde variables"""
    def __init__(self, filepath):
        self.__cdes = {}
        wb = load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=1):
            headers =  [cell.value for cell in row]
        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            metadata = dict(zip(headers, values))
            code = metadata['mip_code']
            mip_type = metadata['mip_type']
            mip_values = metadata['mip_values']
            conceptpath = metadata['conceptPath']
            variable_lookup = metadata['variable_lookup']
            enum_lookup = metadata['enum_lookup']
            cdevar = CdeVariable(code=code, cdetype=mip_type, conceptpath=conceptpath,
                                 mipvalues=mip_values, variable_lookup=variable_lookup,
                                 enum_lookup=enum_lookup)
            self.__cdes[cdevar.code] = cdevar

    @property
    def total_cdes(self):
        return len(self.__cdes)

    def suggest_cde(self, columnreport, threshold=0.6):
        """Suggests the most similar CDE for the column.
        Arguments:
        :param columnreport: ColumnReport object with info of a datset column
        :param threshold: 0-1 similarity threshold, below that not a cde is suggested
        :returns: a CdeVariable object
        """
        name = columnreport.name
        val_range = columnreport.value_range
        mip_type = columnreport.miptype
        LOGGER.debug('The incoming column name is: {}'.format(name))
        # select cdes with tha same type and calculate similarity
        canditates = [cde for cde in self.__cdes.values() if cde.miptype == mip_type]
        LOGGER.debug('Number of cdes with miptype {} is: {}'.format(mip_type, len(canditates)))
        if canditates:
            canditates.sort(key=lambda x: x.similarity(name, val_range), reverse=True)
            canditate = canditates[0]
            similarity = canditate.similarity(name, val_range)
            LOGGER.debug('The simirarity between "{}" and cde "{}" is: {}'.format(name, canditate.code, similarity))
            if similarity >= threshold:            
                return canditate
            else:
                LOGGER.info('No cde match found for incoming column "{}"'. format(name))
                return None
        else:
            LOGGER.info('No cde match found for incoming column "{}"'. format(name))
            return None

    def suggest_replecements(self, cdecode, columnreport, threshold=0.6) -> list:
        """Suggest value replecements for a column for a given cde.
        cde and column must have nominal miptype.
        Arguments:
        :param cdecode: cde code (str)
        :param columnreport: ColumnReport object with info of a datset column
        :returns: List of named Replacement named tupples.
        """
        # get cdevariable object for the given cde code
        cdevar = self.__cdes.get(cdecode)
        replacements = []
        if cdevar:
            # make suggestions only for cases where cde and source col are nominal and also
            # the cdedictionary has values for mipvalues and enumeration lookup columns
            if cdevar.miptype == 'nominal' and columnreport.miptype == 'nominal' and cdevar.mipvalues and cdevar.enum_lookup:
                src_cat = columnreport.value_range
                for enum in src_cat:
                    sug_val = cdevar.suggest_value(enum, threshold)
                    if sug_val:
                        replacements.append(Replacement(enum, sug_val))
                return replacements
            else:
                return None
        else:
            msg = 'Cde "{}" not found in the CdeDictionary'.format(cdecode)
            LOGGER.error(msg)
            raise CdeDictError(msg)


class CdeVariable(object):
    """class for a CDE Variable.
    Arguments:
    :param code: (str) the code of the variable
    :param miptype: (str) the datatype
    :param conceptpath: (str) the concept path 
    :param mipvalues: (str) 'min-max'  for numerical or '{val1, desc1},{val2, desc2},..'
                      for enumerations 
    :param varable_lookup: str with of all the variable names alternatives 'alt1, alt2, alt3...'
    :param enum_lookup: str '{alt1, alt11, alt12...}, {alt21, alt22, alt23..},..
    """
    def __init__(self, code, cdetype, conceptpath,
                 mipvalues=None, variable_lookup=None, enum_lookup=None):
        self.__enum_regx = r'{(?P<list>[^\{]*)}'
        self.__code = code.strip()
        self.__conceptpath = conceptpath.strip()

        # Translating miptype from dictionary file to QC miptype
        cde_type= cdetype.lower().strip()
        if cde_type in QC_FROM_CDE_INTEGER:
            self.__miptype = 'integer'
        elif cde_type in QC_FROM_CDE_NUMERICAL:
            self.__miptype = 'numerical'
        elif cde_type in QC_FROM_CDE_NOMINAL:
            self.__miptype = 'nominal'
        elif cde_type in QC_FROM_CDE_DATE:
            self.__miptype = 'date'
        else:
            self.__miptype = 'text'

        if self.__miptype in ['integer', 'numerical']:
            self.__datatype = 'arithmetic'
        elif self.__miptype == 'nominal':
            self.__datatype = 'nominal'
        else:
            self.__datatype = 'other'

        if variable_lookup:
            # remove double quoues 
            without_quotes = variable_lookup.replace('\"', '')
            lookups = without_quotes.split(',')
            # remove left and right spaces and convert alternative variable names to lowercase
            # remove dublicates and sort them
            self.__variable_lookup = list(set([name.strip().lower() for name in lookups]))
            self.__variable_lookup.sort()
        else:
            self.__variable_lookup = None

        if mipvalues:
            #conver mipvalues to list
            self.__mipvalues = self.__tolist_mipvalues(mipvalues)
        else:
            self.__mipvalues = None
       
        if enum_lookup and mipvalues:
            #convert enum look up to list
            self.__enum_lookup = self.__tolist_enum_lookups(enum_lookup)
            self.__enum_dict = self.__to_dict_enums(enum_lookup)
        else:
            self.__enum_lookup = None
            self.__enum_dict = None

        

    @property
    def code(self):
        return self.__code
    
    @property
    def conceptpath(self):
        return self.__conceptpath
    
    @property
    def miptype(self):
        return self.__miptype

    @property
    def enum_lookup(self):
        return self.__enum_lookup
    
    @property
    def mipvalues(self):
        return self.__mipvalues

    @property
    def variable_lookup(self):
        return self.__variable_lookup

    def similarity(self, name, valrange=None):
        """Returns a similarity score.
        :param name: str with a column name of an incoming dataset
        :param valrange: list with enumerations 
                         or [minimum, maximum] for numerical values
        :return type: float
        """
        namescore = self.__calc_name_f1_score(name)
        if valrange:
            rangescore = self.__calc_range_f1_score(valrange)
            return 0.8 * namescore + 0.2 * rangescore
    
        else:
            return namescore

    def suggest_value(self, value, threshold):
        """Returns the mipmap value if there is match on enum lookups.
        Arguments:
        :param value: (str) incoming categorical value
        :param threshold: 0-1 similarity threshold, below that not a cde is suggested
        """
        if self.__enum_dict and self.__mipvalues:
            lower = value.lower()
            l = {cdeval: max([edit_distance_f1(enum, lower) for enum in enums]) for cdeval, enums in self.__enum_dict.items()}
            suggestion = max(l, key=l.get)
            if l[suggestion] >= threshold:
                return suggestion
            else:
                return None
        else:
            return None
        
    
    def __calc_name_f1_score(self, name) -> float:
        """Returns the best f1 score among lookups for the given name.
        """
        low_name = name.lower()
        from_code = edit_distance_f1(low_name, self.code.lower())
        if self.__variable_lookup:
            dists = [edit_distance_f1(low_name, x.lower()) for x in self.__variable_lookup]
            dists.sort(reverse=True)
            #LOGGER.debug('The edit_distances for cde "{}" and incoming "{}" are: {}'.format(self.code, name, dists))
            from_lookup = dists[0]
            return max(from_code, from_lookup)
        else:
            return from_code

    def __calc_range_f1_score(self, valrange) -> float:
        """Returns a f1 score.
        :param valrange: list with enumerations or list with minimum and maximum
        :type varalange: list with strings or list of float/int 
        :return type: float  
        """
        if self.__datatype == 'nominal' and self.__mipvalues:
            total_enum = len(self.__mipvalues)
            #LOGGER.debug('The cde "{}" in the dictionary has {} enums.'.format(self.code, total_enum))
            # check if there are any enum lookups, otherwise use the enums in mipvalues
            if self.__enum_lookup:
                enums = [x.lower() for x in self.__enum_lookup]
            else:
                enums = [x.lower() for x in self.__mipvalues]
            # convert to lowercase and put all enums lookups in a single list 
            found = sum([str(elem).lower() in enums for elem in valrange])
            incomming_total = len(valrange)
            precision = found / total_enum
            recall = found / incomming_total
            try:
                return 2 * (precision * recall) / (precision + recall)
            except ZeroDivisionError:
                return 0

        elif self.__datatype == 'arithmetic' and self.__mipvalues:
            cde_min = self.__mipvalues[0]
            cde_max = self.__mipvalues[1]
            cde_magnitute = cde_max - cde_min
            minimum, maximum = valrange[0], valrange[1]
            incom_magnitute = maximum - minimum
            if maximum <= cde_min or minimum >= cde_max:
                # incoming val range outside the cde val range            
                inside = 0
            elif minimum <= cde_min:
                if maximum <= cde_max:
                    # there is an overlap on the left side of the cde range
                    inside = maximum - cde_min
                elif maximum >= cde_max:
                    # the incoming range includes the cde range
                    inside = cde_magnitute
            elif minimum >= cde_min:
                if maximum <= cde_max:
                    # the cde range includes the incoming range
                    inside = incom_magnitute
                elif maximum >= cde_max:
                    # there is an overlap on the right side of the cde range
                    inside = cde_max - minimum
            precision = inside / cde_magnitute
            recall = inside / incom_magnitute
            try:
                return 2 * (precision * recall) / (precision + recall)
            except ZeroDivisionError:
                return 0
        else:
            return 0

    def __tolist_enum_lookups(self, enumstr) -> list:
        all_enum_lookups = []
        matches = re.findall(self.__enum_regx, enumstr)
        for m in matches:
            # remove double quates
            without_quotes = m.replace('\"','')
            enums = without_quotes.split(',')
            # remove left and right spaces from each enum and convert to lowercase
            enums = [x.lower().strip() for x in enums]
            all_enum_lookups.extend(enums)
        # remove dublicates and sort
        unique_enums = list(set(all_enum_lookups))
        unique_enums.sort()
        return unique_enums

    def __to_dict_enums(self, enumstr) -> dict:
        all_enum_lookups = []
        matches = re.findall(self.__enum_regx, enumstr)
        for m in matches:
            # remove double quates
            without_quotes = m.replace('\"','')
            enums = without_quotes.split(',')
            # remove left and right spaces from each enum and convert to lowercase
            enums = [x.strip() for x in enums]
            all_enum_lookups.append(enums)
        result = dict(zip(self.mipvalues, all_enum_lookups))
        return result


    def __tolist_mipvalues(self, valuesstr) -> list:
        if self.__datatype == 'arithmetic':
            mipvalues = valuesstr.split('-')
            try:
                mipvalues = [float(x) for x in mipvalues]
                # take the first two elements in case there are more
                mipvalues = mipvalues[:2]
            except ValueError:
                mipvalues = None
            
        elif self.__datatype == 'nominal':
            mipvalues = []
            matches = re.findall(self.__enum_regx, valuesstr)
            for m in matches:
                # remove double quates
                clean_m = m.replace('\"','')
                items = clean_m.split(',')
                # take the fist item from {code, desc}
                enum = items[0]
                mipvalues.append(enum.strip())
            # clean dublicates and sort
            #mipvalues = list(set(mipvalues))
            #mipvalues.sort()
        else:
            mipvalues = None
            
        return mipvalues




