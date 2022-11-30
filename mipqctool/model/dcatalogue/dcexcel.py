import re
import json
from openpyxl import load_workbook

from mipqctool.config import SQL_KEYWORDS, ERROR, LOGGER
from mipqctool.model.qctypes.numerical import infer_numerical

class DcExcel(object):
    """Class for validating a Data Catalogues excel file and optionally produce a dc json."""
    def __init__(self, filepath):
        self.__mand_headers = set([
            'csvfile',
            'name',
            'code',
            'type',
            'values',
            'unit',
            'description',
            'comments',
            'conceptpath',
            'methodology',
        ])
        self.__variables = {}
        self.__conncept_paths = []
        self.__doublicates_cpaths = {}
        self.__invalid_enums = {}
        self.__validation_errors = {}
        self.__invalid_names = []
        wb = load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=1):
            headers =  [cell.value.lower() for cell in row]
        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            rowdata = dict(zip(headers, values))
            code = rowdata['code']
            conceptpath = rowdata['conceptpath']

            if conceptpath not in self.__conncept_paths:
                self.__conncept_paths.append(conceptpath)
            else:
                self.__doublicates_cpaths[code] = conceptpath

            excelvar = ExcelVariable(rowdata)
            if excelvar.invalid_enums:
                self.__invalid_enums[code] = excelvar.invalid_enums
            if excelvar.invalid_code:
                self.__invalid_names.append(excelvar.code)
            if len(excelvar.errors) > 0:
                self.__validation_errors[code] = excelvar.errors

            self.__variables[code] = excelvar


    @property
    def variables(self) -> dict:
        """Dict with variable data"""
        return self.__variables

    @property
    def variable_list(self) -> list:
        """list with variable data"""        
        return list(self.__variables.values())

    @property
    def invalid_enums(self) -> dict:
        return self.__invalid_enums

    @property
    def invalid_names(self) -> list:
        return self.__invalid_names 

    @property
    def validation_errors(self) -> dict:
        return self.__validation_errors

    @property
    def doubl_conceptpaths(self) -> dict:
        return self.__doublicates_cpaths

    @property
    def is_valid(self):
        if (len(self.__validation_errors)) == 0:
            return True
        else:
            return False

    def create_dc_json(self, dataset_code, dataset_name, dataset_version):
        main_group = DcGroup(dataset_code, dataset_name, self.variable_list,  parent=None, version=dataset_version)
        return json.dumps(main_group.group_dict)


class ExcelVariable(object):
    """Class for validating a DC variable from excel file."""
    def __init__(self, rowdata):
        """
        Arguments:
        :param rowdata: dict with variable data stored in in dcexcel row
        """
        self.__variable_dict = None
        self.__errors = []
        self.__invalid_code = False
        self.__enum_regx = '{(?P<list>[^\{]*)}'
        self.__code = rowdata.get('code')
        if not self.__code or self.__code == '':
            self.__errors.append('Variable does not have a code')
        # if name is empty use code as label
        self.__label = rowdata.get('name', self.__code)
        self.__type = None
        self.__sql_type = None
        self.__iscategorical = False
        self.__enums = None
        self.__minValue = None
        self.__maxValue = None
        self.__strvalues = rowdata.get('values')
        self.__conceptpath = rowdata.get('conceptpath')
        self.__units = rowdata.get('unit')
        self.__description = rowdata.get('description')
        self.__comments = rowdata.get('comments')
        self.__methodology = rowdata.get('methodology')
        
        self.__invalid_enums = []
        self.__validate_types(rowdata.get('type', ''))
        self.__validate_concepthpath()
        self.__validate_code()
        if self.__type == 'nominal':
            self.__validate_enums()
            self.__iscategorical = True
        elif self.__type in ['real', 'integer']:
            self.__validate_range()
        
        if len(self.__errors) == 0:
            self.__parent = self.__find_parent()
            self.__recursive_parent = self.__parent
            self.__variable_dict = self.__create_var_dict()
        else:
            self.__parent = None

    @property
    def code(self):
        return self.__code

    @property
    def type(self):
        return self.__type

    @property
    def concepthpath(self):
        return self.__conceptpath

    @property
    def invalid_enums(self):
        return self.__invalid_enums

    @property
    def invalid_code(self):
        return self.__invalid_code
    
    @property
    def errors(self):
        return self.__errors

    @property
    def parent_conceptpath(self):
        return self.__parent

    @property
    def recursive_conceptpath(self):
        LOGGER.debug('Processing variable : {}'.format(self.__code))
        return self.__recursive_parent

    @property
    def var_dict(self):
        return self.__variable_dict

    def update_rconceptpath(self, grandparentcpath):
        self.__recursive_parent = self.__recursive_parent[len(grandparentcpath):]

    def __str__(self) -> str:
        return self.__conceptpath + '/' + self.__code

    def __create_var_dict(self):
        var_dict = {'code': self.__code, 
                    'label': self.__label,
                    'type': self.__type,
                    'sql_type': self.__sql_type,
                    'isCategorical': self.__iscategorical,
        }
        if self.__description:
            var_dict['description'] = self.__description
        else:
            var_dict['description'] = ''
        
        if self.__units:
            var_dict['units'] = self.__units
        else:
            var_dict['units'] = ''
       
        if self.__methodology:
            var_dict['methodology'] = self.__methodology
        else:
            var_dict['methodology'] = ''

        if self.__comments:
            var_dict['comments'] = self.__comments
        else:
            var_dict['comments'] = ''
        
        if self.__type == 'nominal':
            var_dict['enumerations'] = self.__get_enums()
        elif self.__type in ['real', 'integer']:
            if self.__minValue:
                var_dict['minValue'] = self.__minValue
            if self.__maxValue:
                var_dict['maxValue'] = self.__maxValue

        return var_dict


    def __validate_enums(self):
        enumerations = self.__get_enums_codes(self.__strvalues)
        for enum in enumerations:
            if self.__is_valid_enum(enum):
                continue
            else:
                self.__invalid_enums.append(enum)
                self.__errors.append('Variable has invalid enumerations')

    def __get_enums_codes(self, enumstr) -> set:
        all_enums_codes= []
        # list of pairs
        matches = re.findall(self.__enum_regx, enumstr)
        for m in matches:
            items = re.findall(r'"[^"]*"', m)

            # as code we assume that is the fist item of the list
            # the second is the label
            # remove the double quotes
            code = items[0].replace('\"','')
            #remove left and right spaces from each enum and convert to lowercase   
            all_enums_codes.append(code.upper().strip())
        return set(all_enums_codes)

    def __get_enums(self) -> list:
        all_enums = []
        # list of pairs
        matches = re.findall(self.__enum_regx, self.__strvalues)
        for m in matches:
            items = re.findall(r'"[^"]*"', m)
            # remove the double quotes
            try:
                enum_pair = {
                    'code': items[0].replace('\"','').strip(),
                    'label': items[1].replace('\"','').strip()
                }
                #remove left and right spaces from each enum and convert to lowercase   
                all_enums.append(enum_pair)
            except IndexError:
                error = u"There is an invalid character in enumerations ie \u201c, \u201f or \u201d"
                self.__errors.append(error)
        return all_enums

    def __is_valid_enum(self, value):
        if value in SQL_KEYWORDS:
            return False
        # Check if the value start with number
        elif value[:1].isdigit() and not value.isnumeric() :
            return False
        else:
            return True

    def __validate_concepthpath(self):
        try:
            _list = self.concepthpath.split('/')
            variable_code = _list[-1]
            root_group =_list[0]
            if variable_code != self.code:
                self.__errors.append('Concept Path variable name is not the same with variable code')
            if root_group != '':
                #add missing / symbol at the beginning
                self.__conceptpath = '/' + self.__conceptpath
        except AttributeError:
            self.__errors.append('Missing Concept Path or Concept Path is not string')

    def __validate_range(self):
        """validates only positive ranges"""
        if not self.__strvalues:
            pass
        else:
            _r = self.__strvalues.split('-')
            if len(_r) > 2:
                self.__errors.append('Multiple min-max ranges found. Only ranges with possitive numbers supported')
            elif len(_r) == 1:
                self.__errors.append('Single min-max ranges are not supported.')
            else:
                self.__minValue= _r[0]
                self.__maxValue = _r[1]

    def __validate_code(self):
        regex_patern = r'^[a-zA-Z_][a-zA-Z0-9_]*$'
        if re.match(regex_patern, self.code):
            self.__invalid_code = False
        else:
            self.__errors.append("Variable {} has invalid name.".format(self.__code))
            self.__invalid_code = True


    def __find_parent(self):
        return self.__conceptpath[ :-(len(self.__code) +1)]

    def __validate_types(self, typecode):
        if typecode == 'nominal':
            self.__type = typecode
            self.__sql_type = 'text'
        elif typecode == 'ordinal':
            self.__type = typecode
            self.__sql_type = 'text'
        elif typecode == 'text':
            self.__type = typecode
            self.__sql_type = 'text'
        elif typecode == 'integer':
            self.__type = typecode
            self.__sql_type = 'int'
        elif typecode == 'real':
            self.__type = typecode
            self.__sql_type = 'real'
        elif typecode == 'date':
            self.__type = typecode
            self.__sql_type = 'date'
        else:
            self.__errors.append('Invalid variable data type') 



class DcGroup(object):
    """Class for creating """

    def __init__(self, name, code, variables, parent=None, version=None):
        """
        :param name: string with the conceptpath string
        :param variables: list with ExcelVariable objects
        :param parent: a DcGroup object
        """
        self.__variables = []
        self.__groups = []
        self.__children_variables = {}  
        self.__code = code
        self.__label = name
        self.__parent = parent
        self.__conceptpath = self.__create_conceptpath()
        self.__group_dict = None
        self.__version = version
        LOGGER.debug('Processing node {}'.format(self.__code))
        
        for variable in variables:
            self.__categorize_variable(variable)
        
        LOGGER.debug('Current node {} has variables: {}'.format(self.__code, self.__variables))
        LOGGER.debug('Current node {} has childs: {}'.format(self.__code, self.__children_variables.items()))

        if len(self.__children_variables) > 0:
            for children_name, cvariables in self.__children_variables.items():
                child = DcGroup(name=children_name, code=children_name,
                                variables=cvariables, parent=self)
                self.__groups.append(child)
    
    def __str__(self) -> str:
        return self.__conceptpath


    @property
    def conceptpath(self):
        return self.__conceptpath
    
    @property
    def recursive_conceptpath(self):
        if self.__parent:
            return '/' + self.__code
        else:
            return '/'

    @property
    def group_dict(self):
        _dict = {
            'code': self.__code,
            'label': self.__label
        }
        if not self.__parent:
            _dict['version'] = self.__version
        if len(self.__variables) > 0:
            _dict['variables'] = [variable.var_dict for variable in self.__variables]
        if len(self.__groups) > 0:
            _dict['groups'] = [group.group_dict for group in self.__groups]
        return _dict

    @property
    def variables(self):
        return self.__variables

    @property
    def children_variables(self):
        return self.__children_variables


    def __create_conceptpath(self):
        conceptpath = '/'
        if self.__parent:
            conceptpath = self.__parent.conceptpath +  self.__code + '/'
        return conceptpath

    def __categorize_variable(self, variable):
        LOGGER.debug('Currnet variable {} has recursive cpath: {}'.format(variable.code, variable.recursive_conceptpath))

        if variable.recursive_conceptpath == '':
            self.__variables.append(variable)
        else:
            children_name = variable.recursive_conceptpath.split('/')[1]
            if children_name in self.__children_variables.keys():
                variable.update_rconceptpath('/' + children_name)
                self.__children_variables[children_name].append(variable)
            else:
                variable.update_rconceptpath('/' + children_name)
                self.__children_variables[children_name] = [variable]





