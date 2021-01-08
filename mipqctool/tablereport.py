# -*- coding: utf-8 -*-
# columnreport.py

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function
from __future__ import unicode_literals

import os
import csv
from datetime import datetime

import pandas as pd
from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from collections import namedtuple, Counter, defaultdict

from mipqctool.exceptions import TableReportError, QCToolException
from mipqctool.columnreport import ColumnReport
from mipqctool.qcfrictionless import QcTable
from mipqctool import config, qctypes, __version__
from mipqctool.config import LOGGER, COLUMN_STAT_HEADERS


config.debug(True)


class TableReport(object):
    """This class is for creating a report in pdf and csv files
    """

    def __init__(self, table, id_column=1, threshold=3, **options):
        """ Arguments:
            :param table: a QcTable object
            :param id_column: column number of dataset's primary key (id)
            :param threshold: outlier threshold - (mean - threshold * std, mean + threshold * std) 
                              outside this length, a numerical value is considered outlier
        """
        self.__threshold = threshold
        self.__missing_headers = []
        self.__valid_headers = []
        self.__invalid_headers = []

        # check if table has a schema else infer
        if not table.schema:
            table.infer()

        # check if id_column number exist and get its column name
        try:

            self.__id_column = table.schema.fields_names[id_column-1]
            self.__id_index = id_column - 1

        # id column does not exist
        except IndexError:
            self.__id_column = None
            self.__id_index = None
            raise QCToolException("Could not find any columns in the csv. Please check the seperator.")


        self.__table = table

        self.__total_columns = len(self.__table.schema.field_names)
        if self.__table.with_metadata:
            self.__validate_headers()
        self.__columns_quantiles = None
        # Calc which column number corresponds to which quantile
        self.__calc_columns_quantiles()
        # list to hold ColumnReport objects
        self.__columnreports = []

        self.__total_rows = None
        self.__rows_only_id = None
        self.__rows_no_id = None
        self.__tvalid_columns = None
        self.__tfilled_columns = None
        self.__corrected = False

        self.__create_reports()
        self.__collect_row_stats()

    @property
    def table(self):
        """Return the QcTable object"""
        return self.__table
    
    @property
    def columnreports(self):
        """A list of ColumnReport objects."""
        return self.__columnreports
    
    @property
    def total_rows(self):
        """Number of dataset's rows"""
        return self.__total_rows

    @property
    def corrected(self):
        """Are the correction applied?"""
        return self.__corrected

    @property
    def total_columns(self):
        """Number of dataset's columns"""
        return self.__total_columns

    @property
    def valid_rows_stats(self):
        """Dict with rows data validation overall stats"""
        return self.__valid_rows_stats

    @property
    def filled_rows_stats(self):
        """Dict with rows data completion overall stats"""
        return self.__filled_rows_stats

    @property
    def missing_headers(self):
        return self.__missing_headers

    @property
    def valid_headers(self):
        return self.__valid_headers

    @property
    def invalid_headers(self):
        return self.__invalid_headers

    @property
    def isvalid(self):
        all_invalid_headers = self.invalid_headers + self.missing_headers
        if all_invalid_headers or self.__total_invalid_rows != 0:
            return False
        else:
            return True

    def apply_corrections(self):
        """Applies the suggestions of invalid values to the dataset."""
        for columnreport in self.__columnreports:
            columnreport.apply_corrections()
        self.__collect_row_stats()
        self.__corrected = True

    def save_corrected(self, path):
        if self.__corrected:
            list_new_values = (col.corrected_values for col in self.__columnreports)
            new_values = zip(*list_new_values)
            with open(path, 'w') as out:
                csv_out = csv.writer(out, quoting=csv.QUOTE_ALL)
                csv_out.writerow(self.__table.headers)
                for row in new_values:
                    csv_out.writerow(row)

    def printpdf(self, filepath):
        app_path = os.path.abspath(os.path.dirname(__file__))
        env_path = os.path.join(app_path, 'html')
        css_path = os.path.join(env_path, 'style.css')
        docs = []
        template_vars = self.__prepare_stat2pdf()
        env = Environment(loader=FileSystemLoader(env_path))
        if self.__corrected:
            template_file = 'dataset_report_verified.html'
        else:
            template_file = 'dataset_report.html'
        template = env.get_template(template_file)
        html_out = template.render(template_vars)
        docs.append(HTML(string=html_out).render(stylesheets=[css_path]))
        for columnreport in self.__columnreports:
            docs.append(HTML(string=columnreport.to_html()).render(stylesheets=[css_path]))

        all_pages = [p for doc in docs for p in doc.pages]
        docs[0].copy(all_pages).write_pdf(target=filepath)

    def printexcel(self, filepath):
        # create an excel workbook
        wb = Workbook()

        wb1 = wb.active

        ## spreadsheet "General"##
        wb1.title = 'General'
        col = wb1.column_dimensions['A']
        col.font = Font(bold=True)
        col.width = 35

        daterun = datetime.now()
        # use pretty values for boolean
        if self.__corrected:
            applied_corrections = 'Yes'
        else:
            applied_corrections = 'No'

        if self.__table.with_metadata:
            use_metadata = 'Yes'
        else:
            use_metadata = 'No'
        # start filling the sheet "General"
        wb1.append(['Dataset file', self.__table.filename])
        wb1.append(['Dataset filepath', self.__table.source])
        wb1.append(['QC tool version', __version__])
        wb1.append(['Date qc run', daterun.strftime("%d/%m/%Y %H:%M:%S")])
        wb1.append(['Total columns', self.total_columns])
        wb1.append(['Total rows', self.total_rows])
        wb1.append(['Metadata used', use_metadata])
        wb1.append(['Missing columns'] + self.missing_headers)
        wb1.append(['Extra columns'] + self.invalid_headers)
        wb1.append(['Invalid rows', self.__total_invalid_rows])

        ws2 = wb.create_sheet("Row Statistics")
        col2 = ws2.column_dimensions['A']
        col2.width = 40
        col2.font = Font(bold=True)
        #ws2.append(['rows with only id column filled', len(self.__rows_only_id)])
        #ws2.append(['rows with no id column filled', len(self.__rows_no_id)])
        ws2.append(['rows with 0-24% of the columns filled', self.filled_rows_stats['filled_0_24']])
        ws2.append(['rows with 25-49% of the columns filled', self.filled_rows_stats['filled_25_49']])
        ws2.append(['rows with 50-74% of the columns filled', self.filled_rows_stats['filled_50_74']])
        ws2.append(['rows with 75-99% of the columns filled', self.filled_rows_stats['filled_75_99']])
        ws2.append(['rows with 100% of the columns filled', self.filled_rows_stats['filled_100']])

        ws2.append(['rows with 0-24% of the columns valid', self.valid_rows_stats['valid_0_24']])
        ws2.append(['rows with 25-49% of the columns valid', self.valid_rows_stats['valid_25_49']])
        ws2.append(['rows with 50-74% of the columns valid', self.valid_rows_stats['valid_50_74']])
        ws2.append(['rows with 75-99% of the columns valid', self.valid_rows_stats['valid_75_99']])
        ws2.append(['rows with 100% of the columns valid', self.valid_rows_stats['valid_100']])
        
        chart1 = BarChart()
        chart1.type = 'bar'
        chart1.style = 11
        chart1.y_axis.title = '# of rows'
        chart1.title = 'Number of rows per filled columns'
        values1 = Reference(ws2, min_col=1, min_row=3, max_row=7, max_col=2)
        chart1.add_data(values1)
        chart1.shape = 4
        ws2.add_chart(chart1, 'D1')

        chart2 = BarChart()
        chart2.type = 'bar'
        chart2.style = 12
        chart2.y_axis.title = '# of rows'
        chart2.title = 'Number of rows per valid columns'
        values2 = Reference(ws2, min_col=1, min_row=8, max_row=13, max_col=2)
        chart2.add_data(values2)
        chart2.shame = 4
        ws2.add_chart(chart2, 'D20')

        ## Column Statistics Sheet ## 
        ws3 = wb.create_sheet("Column Statistics")
        # make bold the first column 
        title_row = ws3.row_dimensions[1]
        title_row.font = Font(bold=True)
        # get the columns stats and fill the rows
        df_cstats = self.__column_stats_2_df()       
        for r in dataframe_to_rows(df_cstats, index=False, header=True):
            ws3.append(r)
        for i in range(1, len(df_cstats.columns) + 1):
            ws3.column_dimensions[get_column_letter(i)].width = 20

        # start filling Cleaning suggestions sheet
        if self.corrected:
            cleaning_sheetname = 'Cleaned values'
        else:
            cleaning_sheetname = 'Cleaning suggestions'
        ws4 = wb.create_sheet(cleaning_sheetname)        
        left_border = Border(left=Side(style='thick'))
        right_border = Border(right=Side(style='thick'))
        header_border = Border(left=Side(style='thick'), right=Side(style='thick'))
        center_alignment = Alignment(horizontal='center')                     
        start_col = 1
        for colreport in self.__columnreports:
            end_col = start_col + 1
            ws4.column_dimensions[get_column_letter(start_col)].width = 15
            ws4.column_dimensions[get_column_letter(end_col)].width = 15

            # merge the two cells in the first row and put the variable name as title
            ws4.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            header = ws4.cell(row=1,column=start_col, value=colreport.qcfield.name)
            header.border = header_border
            header.alignment = center_alignment
            header.font = Font(bold=True)
            # add also the subtitlles
            ws4.cell(row=2, column=start_col, value='Invalid Value').border = left_border
            ws4.cell(row=2, column=end_col, value='Corrected Value').border = right_border
            # collect all the correction suggestions for invalid values
            # constraint violations
            corrections = list(colreport.ccorrections)
            ctonulls = [(value, 'Null') for value in colreport.cnulls]
            # datatype violations
            corrections.extend(list(colreport.dcorrections))
            dtonulss = [(value, 'Null') for value in colreport.dnulls]
            # append null sugestions to corrections
            corrections.extend(ctonulls)
            corrections.extend(dtonulss)
            # now start filling the rows for that variable 
            start_row = 3
            for pair in corrections:
                left = ws4.cell(row=start_row, column=start_col, value=pair[0])
                left.border = left_border
                right = ws4.cell(row=start_row, column=end_col, value=pair[1])
                right.border = right_border
                start_row += 1
            start_col += 2       

        wb.save(filepath)

    # private
    def __create_reports(self):
        """Create column reports."""
        for qcfield in self.__table.schema.fields:
            try:
                raw_values = self.__table.column_values(qcfield.name)
                column_report = ColumnReport(raw_values, qcfield, threshold=self.__threshold)
                column_report.validate()
                self.__columnreports.append(column_report)
            except QCToolException:
                pass

    def __collect_row_stats(self):
        # get the rows with no id
        # get the report of id column
        id_column_report = self.__columnreports[self.__id_index]
        rows_with_no_id = id_column_report.null_row_numbers

        # find rows with only id filled in and the total row number
        # find total nulls and invalid
        rows_with_only_id = id_column_report.filled_row_numbers
        total_rows = id_column_report.total_rows

        rows_invalid = []
        rows_nulls = []
        # for each column
        for report in self.__columnreports:
            rows_invalid.extend(list(report.invalid_rows))
            rows_nulls.extend(list(report.null_row_numbers))
            if report.qcfield.name == self.__id_column:
                continue
            else:
                # substract the row numbers that have filled values 
                # at the end of the loop only the rows with only id filled will remain
                rows_with_only_id = rows_with_only_id - report.filled_row_numbers
                # this if statement cover the case where there are rows with missing id 
                if report.total_rows > total_rows:
                    total_rows = report.total_rows

        total_invalid_rows = len(set(rows_invalid))
        invalid_counter = Counter(rows_invalid)
        null_counter = Counter(rows_nulls)
        max_columns_per_row_ = {row: self.total_columns
                                for row in range(1, total_rows + 1)}
        # initialize valid and filled counters with max column
        # number per row. aka each row has all columns filled and valid
        valid_counter = Counter(max_columns_per_row_)
        filled_counter = Counter(max_columns_per_row_)

        # subtract the invalid to find the valid columns per row
        valid_counter.subtract(invalid_counter)
        # subtract the nulls to find the filled columns per row
        filled_counter.subtract(null_counter)

        self.__total_rows = total_rows
        self.__total_invalid_rows = total_invalid_rows

        self.__rows_only_id = rows_with_no_id
        self.__rows_no_id = rows_with_only_id
        self.__tvalid_columns = self.__calc_rows_per_columns(dict(valid_counter))
        self.__tfilled_columns = self.__calc_rows_per_columns(dict(filled_counter))
        self.__valid_rows_stats = self.__calc_rstat_dict(columns='valid')
        self.__filled_rows_stats = self.__calc_rstat_dict(columns='filled')

    def __calc_columns_quantiles(self):
        """Calcs which column number corresponds to which quantile. """
        cquantiles = {}
        cquantiles['p99'] = self.total_columns - 1
        cquantiles['p75'] = round(self.total_columns * 75 / 100)
        cquantiles['p74'] = cquantiles['p75'] - 1
        cquantiles['p50'] = round(self.total_columns / 2)
        cquantiles['p49'] = cquantiles['p50'] - 1
        cquantiles['p25'] = round(self.total_columns * 25 / 100)
        cquantiles['p24'] = cquantiles['p25'] - 1
        self.__columns_quantiles = cquantiles

    def __calc_rows_per_columns(self, d):
        """
        Arguments:
        :param d: dictionary with row number as keys and values 
                  the number of valid (or filled) columns that this row has
        :return: dictionary with number of valid (or filled ) columns as keys
                 and values the number of rows that have the same 
                 valid (or filled) number of columns 
        """
        res = defaultdict(list)
        for key, val in d.items():
            res[val].append(key)
        result = {k: len(v) for (k, v) in res.items()}
        return result

    def __calc_rstat_dict(self, columns='filled'):
        stats = {}
        if columns == 'filled':
            dict_column = self.__tfilled_columns
        elif columns == 'valid':
            dict_column = self.__tvalid_columns

        for i in range(0, self.__columns_quantiles['p25']):
            key = columns + '_' + '0_24'
            stats[key] = stats.get(key, 0) + dict_column.get(i, 0)

        for i in range(self.__columns_quantiles['p25'], self.__columns_quantiles['p50']):
            key = columns + '_' + '25_49'
            stats[key] = stats.get(key, 0) + dict_column.get(i, 0)

        for i in range(self.__columns_quantiles['p50'], self.__columns_quantiles['p75']):
            key = columns + '_' + '50_74'
            stats[key] = stats.get(key, 0) + dict_column.get(i, 0)

        for i in range(self.__columns_quantiles['p75'], self.__total_columns):
            key = columns + '_' + '75_99'
            stats[key] = stats.get(key, 0) + dict_column.get(i, 0)

        key = columns + '_' + '100'
        stats[key] = dict_column.get(self.total_columns, 0)

        return stats

    def __prepare_stat2pdf(self):
        daterun = datetime.now()

        if self.__corrected:
            applied_corrections = 'Yes'
        else:
            applied_corrections = 'No'

        if self.__table.with_metadata:
            use_metadata = 'Yes'
        else:
            use_metadata = 'No'

        html_vars = {
            'datasetfile': self.__table.filename,
            'date_run': daterun.strftime("%d/%m/%Y %H:%M:%S"),
            'qcversion': __version__,
            'csvfilepath': self.__table.source,
            'use_metadata': use_metadata,
            'applied_corrections': applied_corrections,
            #'only_ids': len(self.__rows_only_id),
            #'only_ids_perc': round(len(self.__rows_only_id) / self.total_rows * 100, 2),
            #'no_ids': len(self.__rows_no_id),
            #'no_ids_perc': round(len(self.__rows_no_id) / self.total_rows * 100, 2),
            'total_columns': self.total_columns,
            'total_rows': self.total_rows,
            'total_invalid_rows': self.__total_invalid_rows
        }

        html_vars.update(self.filled_rows_stats)
        html_vars.update(self.valid_rows_stats)
        html_vars.update(self.__columns_quantiles)

        # calc the percentages for valid and filled row stats
        for item in self.filled_rows_stats.items():
            key_perc = item[0] + '_perc'
            html_vars[key_perc] = round(item[1] / self.total_rows * 100, 2)

        for item in self.valid_rows_stats.items():
            key_perc = item[0] + '_perc'
            html_vars[key_perc] = round(item[1] / self.total_rows * 100, 2)

        return html_vars

    def __validate_headers(self):
        # metadata json is provided so we can validate the headers
        valid_headers = []
        missing_headers = []
        invalid_headers = []
        schema_headers = self.__table.schema.field_names
        actual_headers = self.__table.actual_headers
        for name in schema_headers:
            if name in actual_headers:
                valid_headers.append(name)
            else:
                missing_headers.append(name)
        invalid_headers = [
            header for header in actual_headers
            if header not in schema_headers
        ]

        self.__valid_headers = valid_headers
        self.__missing_headers = missing_headers
        self.__invalid_headers = invalid_headers

    def __column_stats_2_df(self) -> pd.DataFrame:
        d = defaultdict(list)
        for column in self.__columnreports:
            d['name'].append(column.qcfield.name)
            d['type'].append(column.miptype)
            d['filled %'].append(column.filledpercentage)
            invalid_values = column.constraint_errors + column.datatype_errors
            d['invalid values'].append(invalid_values)
            for key in COLUMN_STAT_HEADERS:
                value = column.stats.get(key)
                if value:
                    # convert value to string
                    # list '['cat1', 'cate2', ..]
                    value = str(value)
                elif value == []:
                    value = None
                d[key].append(value)
        df = pd.DataFrame.from_dict(d)
        headers = ['name', 'type', 'filled %', 'invalid values'] + COLUMN_STAT_HEADERS
        return df[headers]
